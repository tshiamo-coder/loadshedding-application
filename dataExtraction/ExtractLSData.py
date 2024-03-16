#Script to extract data out of all the tables in the /data dirctory

import pandas as pd
from openpyxl import Workbook
import openpyxl
import json
import os

#instatntiate workbook object
#testing opnpyxl
# wb = openpyxl.load_workbook("data/WesternCape_LS.xlsx")
# print(wb.sheetnames)
#end of testing ====
def get_Sheet_Data(sheet):
    print(sheet)
    return
#Get excel sheet based on excel filename and sheetname
def get_Sheet_From_File(filename, sheetname):
    workbook =  openpyxl.load_workbook(filename,data_only=True)
    try:
        print(workbook.sheetnames)
        worksheet = workbook[sheetname]
    except:
        print("ERROR%******")
        return
    return worksheet

#Get loadshedding data 
def get_loadshedding_data(ws):
    list_of_schemas = []
    data_schema = {}
    for row in range(16,112):
        print(ws.cell(row,3).value)
        if(ws.cell(row,3).value == 1):
            # refresh data schema 
            data_schema = {
        "start":str(ws.cell(row,1).value),
        "end":str(ws.cell(row,2).value),
        "loadshedding":{
            1:{},
            2:{},
            3:{},
            4:{},
            5:{},
            6:{},
            7:{},
            8:{},
        }
            }
            #start of if indent
            print(f"{ws.cell(row,1).value}-{ws.cell(row,2).value}")
        rowData = ""
        #add start time 
        #add start time for specefic group
        row_list = []
        for col in range (4,35):
            # print(f"=======getting data {row}==========")
            row_list.append(ws.cell(row,col).value)
        #add stage no and groups affected from day 1-31
        data_schema["loadshedding"][ws.cell(row,3).value] = row_list
        print(row_list)
        #when the 8th group is reached start from 1 again
        if ws.cell(row,3).value ==8:
            print(data_schema)
            list_of_schemas.append(data_schema)

        # print(ws.cell(111,3).value)
    return list_of_schemas

def get_all_suburbs(provinces):
    all_suburbs = []
    concat_str = "_LS.xlsx"
    #loop through every province
    for p in provinces:
        filename = "dataExtraction/data/"+p+concat_str
        sheet = get_Sheet_From_File(filename,sheetname="SP_List")
        max_row = sheet.max_row
        province_surburbs = []
        for row in range (1,max_row+1):
            suburb_schema = {
                "SName": sheet.cell(row,4).value, #4
                "MpName": sheet.cell(row,2).value, #2
                "Block": sheet.cell(row,7).value, #7
                "Type": sheet.cell(row,8).value, #8
                "Province": p
            }
            print(suburb_schema)
            province_surburbs.append(suburb_schema)
        all_suburbs.extend(province_surburbs)

    return all_suburbs

if __name__ == "__main__":
    provinces = ["WesternCape","EasternCape","NorthernCape",
                 "NorthWest", "FreeState","Gauteng","Mpumalanga",
                 "Limpopo","KwaZulu-Natal"]
    #file to extract loadshedding tables
    # filename ="dataExtraction/data/WesternCape_LS.xlsx"
    # sheet = get_Sheet_From_File(filename,sheetname="Schedule")
    # get_Sheet_Data(sheet)
    # dict_table = get_loadshedding_data(sheet)
    # tuple_obj = (*dict_table,)
    # # dict_data = json.loads(str(dict_table))
    # jsonObj = json.dumps(dict_table)
    # with open("ls_data.json", "w") as outfile:
    #     outfile.write(jsonObj)

    #THe below functions is for extracting all the suburbs
    all_suburbs = get_all_suburbs(provinces)
    # cwd = os.getcwd()
    # print(cwd)
    # ws = get_Sheet_From_File(filename,"SP_List")
    jsonObj = json.dumps(all_suburbs)
    with open("subrub_data.json", "w") as outfile:
        outfile.write(jsonObj)